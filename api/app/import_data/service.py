from datetime import datetime
from io import BytesIO

import openpyxl
from bson import ObjectId
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.auth.service import list_all_user_summaries
from app.cases.models import CaseStatus, CaseType, ReporterType
from app.cases.schemas import CaseImportCreate
from app.cases.service import create_case_from_import
from app.import_data.schemas import ImportResultOut, ImportRowError
from app.reference_data.service import list_reference_items

# Column order matches plan §13/T13.1 exactly — `CaseId`/`Date of Closure`
# are deliberately excluded, both are server-generated.
COLUMNS = [
    "Reported Date",
    "Reporter Type",
    "Reporter Name",
    "Customer",
    "Product",
    "Category",
    "Description",
    "Assigned To",
    "Status",
    "Type",
    "Market",
    "Remarks",
    "Bug Number",
    "Task Numbers",
    "WO Numbers",
    "Resolution",
]

REPORTER_TYPES: list[str] = list(ReporterType.__args__)  # type: ignore[attr-defined]
STATUSES: list[str] = list(CaseStatus.__args__)  # type: ignore[attr-defined]
TYPES: list[str] = list(CaseType.__args__)  # type: ignore[attr-defined]

TEMPLATE_MAX_ROWS = 2000  # how many blank rows get the dropdown validation


async def build_template_workbook() -> bytes:
    """Plan §13/T13.1-T13.2 — a clean, human-fillable `.xlsx` matching the
    current schema, with dropdown validation on the enum-like columns so a
    filled-in template can't contain a typo'd status/type/product/category."""
    products = [i.value for i in await list_reference_items("product", True)]
    categories = [i.value for i in await list_reference_items("category", True)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cases"

    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for col_idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 22

    # Long/dynamic option lists (Product, Category) can't be inlined as a
    # `"a,b,c"` formula — Excel's inline data-validation list has a ~255
    # char limit — so they live on a hidden "Lists" sheet and are
    # referenced by range instead.
    lists_ws = wb.create_sheet("Lists")
    lists_ws.sheet_state = "hidden"
    _write_list_column(lists_ws, 1, "ReporterType", REPORTER_TYPES)
    _write_list_column(lists_ws, 2, "Status", STATUSES)
    _write_list_column(lists_ws, 3, "Type", TYPES)
    _write_list_column(lists_ws, 4, "Product", products)
    _write_list_column(lists_ws, 5, "Category", categories)

    _add_dropdown(ws, "B", lists_ws, 1, len(REPORTER_TYPES))  # Reporter Type
    _add_dropdown(ws, "E", lists_ws, 4, len(products))  # Product
    _add_dropdown(ws, "F", lists_ws, 5, len(categories))  # Category
    _add_dropdown(ws, "I", lists_ws, 2, len(STATUSES))  # Status
    _add_dropdown(ws, "J", lists_ws, 3, len(TYPES))  # Type

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_list_column(ws, col_idx: int, header: str, values: list[str]) -> None:
    ws.cell(row=1, column=col_idx, value=header)
    for i, value in enumerate(values, start=2):
        ws.cell(row=i, column=col_idx, value=value)


def _add_dropdown(ws, target_col_letter: str, lists_ws, lists_col_idx: int, count: int) -> None:
    if count == 0:
        return
    lists_col_letter = get_column_letter(lists_col_idx)
    formula = f"Lists!${lists_col_letter}$2:${lists_col_letter}${count + 1}"
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{target_col_letter}2:{target_col_letter}{TEMPLATE_MAX_ROWS}")


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _parse_date(value) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str) and value.strip():
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
    return None


def _split_list_cell(value) -> list[str]:
    text = _clean(value)
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part.strip()]


async def parse_and_import(file_bytes: bytes, current_user_id: ObjectId) -> ImportResultOut:
    """Plan §13/T13.3, T13.6 — reject bad rows with a clear per-row reason
    rather than silently skipping (unlike Phase 5's `migrate_excel.py`)."""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb["Cases"] if "Cases" in wb.sheetnames else wb.active

    users_by_lower_name = {u.name.strip().lower(): u for u in await list_all_user_summaries()}

    total_rows = 0
    imported = 0
    rejected: list[ImportRowError] = []

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = list(row) + [None] * (len(COLUMNS) - len(row))
        (
            reported_date_raw,
            reporter_type_raw,
            reporter_name_raw,
            customer_raw,
            product_raw,
            category_raw,
            description_raw,
            assigned_to_raw,
            status_raw,
            type_raw,
            market_raw,
            remarks_raw,
            bug_number_raw,
            task_numbers_raw,
            wo_numbers_raw,
            resolution_raw,
        ) = cells[: len(COLUMNS)]

        if all(c is None for c in cells[: len(COLUMNS)]):
            continue  # blank trailing row — not a real data row
        total_rows += 1

        reported_date = _parse_date(reported_date_raw)
        reporter_type = _clean(reporter_type_raw)
        reporter_name = _clean(reporter_name_raw)
        customer = _clean(customer_raw)
        product = _clean(product_raw)
        category = _clean(category_raw)
        description = _clean(description_raw)
        assigned_to_name = _clean(assigned_to_raw)
        status = _clean(status_raw)
        case_type = _clean(type_raw)

        missing = [
            name
            for name, value in [
                ("Reported Date", reported_date),
                ("Reporter Type", reporter_type),
                ("Reporter Name", reporter_name),
                ("Customer", customer),
                ("Product", product),
                ("Category", category),
                ("Description", description),
                ("Assigned To", assigned_to_name),
                ("Status", status),
                ("Type", case_type),
            ]
            if value is None
        ]
        if missing:
            rejected.append(
                ImportRowError(row=row_number, reason=f"Missing required field(s): {', '.join(missing)}")
            )
            continue

        if reporter_type not in REPORTER_TYPES:
            rejected.append(
                ImportRowError(
                    row=row_number,
                    reason=f"Invalid Reporter Type '{reporter_type}' — must be one of {REPORTER_TYPES}",
                )
            )
            continue
        if status not in STATUSES:
            rejected.append(
                ImportRowError(row=row_number, reason=f"Invalid Status '{status}' — must be one of {STATUSES}")
            )
            continue
        if case_type not in TYPES:
            rejected.append(
                ImportRowError(row=row_number, reason=f"Invalid Type '{case_type}' — must be one of {TYPES}")
            )
            continue

        assignee = users_by_lower_name.get(assigned_to_name.lower())
        if assignee is None:
            rejected.append(
                ImportRowError(
                    row=row_number,
                    reason=f"Assigned To '{assigned_to_name}' does not match any existing user",
                )
            )
            continue

        payload = CaseImportCreate(
            reportedDate=reported_date,
            reporterType=reporter_type,
            reporterName=reporter_name,
            customer=customer,
            product=product,
            category=category,
            description=description,
            assignedTo=assignee.id,
            status=status,
            type=case_type,
            market=_clean(market_raw) or "",
            remarks=_clean(remarks_raw) or "",
            resolution=_clean(resolution_raw) or "",
            bugNumber=_clean(bug_number_raw),
            taskNumbers=_split_list_cell(task_numbers_raw),
            workOrderNumbers=_split_list_cell(wo_numbers_raw),
        )
        await create_case_from_import(payload, current_user_id)
        imported += 1

    return ImportResultOut(totalRows=total_rows, imported=imported, rejected=rejected)
