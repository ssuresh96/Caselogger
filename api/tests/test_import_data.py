from datetime import datetime
from io import BytesIO

import openpyxl
from bson import ObjectId
from httpx import AsyncClient

from app.auth.models import User
from app.core.database import database
from app.import_data.service import COLUMNS


def _auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _build_xlsx(rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cases"
    for col_idx, header in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _valid_row(**overrides) -> list:
    row = {
        "reported_date": datetime(2025, 6, 1),
        "reporter_type": "Customer",
        "reporter_name": "Jane Reporter",
        "customer": "Acme Corp",
        "product": "AOS",
        "category": "AOS-General Support",
        "description": "Historical case from before the app existed",
        "assigned_to": "Test Agent",
        "status": "Resolved",
        "type": "Support",
        "market": "UAE",
        "remarks": "Handled over email",
        "bug_number": "BUG-1",
        "task_numbers": "TASK-1, TASK-2",
        "wo_numbers": "WO-1",
        "resolution": "Fixed config",
    }
    row.update(overrides)
    return [
        row["reported_date"],
        row["reporter_type"],
        row["reporter_name"],
        row["customer"],
        row["product"],
        row["category"],
        row["description"],
        row["assigned_to"],
        row["status"],
        row["type"],
        row["market"],
        row["remarks"],
        row["bug_number"],
        row["task_numbers"],
        row["wo_numbers"],
        row["resolution"],
    ]


async def test_template_download_requires_superuser(
    client: AsyncClient, regular_user_token: str
):
    resp = await client.get("/admin/import/cases/template", headers=_auth(regular_user_token))
    assert resp.status_code == 403


async def test_template_download_requires_auth(client: AsyncClient):
    resp = await client.get("/admin/import/cases/template")
    assert resp.status_code == 401


async def test_template_download_returns_valid_workbook(
    client: AsyncClient, superuser_token: str
):
    resp = await client.get("/admin/import/cases/template", headers=_auth(superuser_token))
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws = wb["Cases"]
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header_row == COLUMNS


async def test_import_requires_superuser(client: AsyncClient, regular_user_token: str):
    xlsx = _build_xlsx([_valid_row()])
    resp = await client.post(
        "/admin/import/cases",
        headers=_auth(regular_user_token),
        files={"file": ("import.xlsx", xlsx, "application/octet-stream")},
    )
    assert resp.status_code == 403


async def test_import_valid_rows_creates_cases(
    client: AsyncClient, superuser_token: str, superuser: User, regular_user: User
):
    xlsx = _build_xlsx([_valid_row(), _valid_row(customer="Second Customer")])
    resp = await client.post(
        "/admin/import/cases",
        headers=_auth(superuser_token),
        files={"file": ("import.xlsx", xlsx, "application/octet-stream")},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["totalRows"] == 2
    assert body["imported"] == 2
    assert body["rejected"] == []

    resp = await client.get(
        "/cases", params={"search": "Acme Corp"}, headers=_auth(superuser_token)
    )
    items = resp.json()["items"]
    assert len(items) == 1
    case = items[0]
    assert case["assignedTo"]["id"] == str(regular_user.id)
    assert case["bugNumber"] == "BUG-1"
    assert case["taskNumbers"] == ["TASK-1", "TASK-2"]
    assert case["workOrderNumbers"] == ["WO-1"]
    assert case["market"] == "UAE"
    assert case["remarks"] == "Handled over email"
    assert case["resolution"] == "Fixed config"

    entries = await database["activity_log"].find({"case_id": ObjectId(case["id"])}).to_list(None)
    assert len(entries) == 1
    assert "imported" in entries[0]["change_summary"]


async def test_import_rejects_missing_required_field(
    client: AsyncClient, superuser_token: str
):
    xlsx = _build_xlsx([_valid_row(customer=None)])
    resp = await client.post(
        "/admin/import/cases",
        headers=_auth(superuser_token),
        files={"file": ("import.xlsx", xlsx, "application/octet-stream")},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["imported"] == 0
    assert len(body["rejected"]) == 1
    assert body["rejected"][0]["row"] == 2
    assert "Customer" in body["rejected"][0]["reason"]


async def test_import_rejects_invalid_status(client: AsyncClient, superuser_token: str):
    xlsx = _build_xlsx([_valid_row(status="NotAStatus")])
    resp = await client.post(
        "/admin/import/cases",
        headers=_auth(superuser_token),
        files={"file": ("import.xlsx", xlsx, "application/octet-stream")},
    )
    body = resp.json()
    assert body["imported"] == 0
    assert "Invalid Status" in body["rejected"][0]["reason"]


async def test_import_rejects_unmatched_assignee(client: AsyncClient, superuser_token: str):
    xlsx = _build_xlsx([_valid_row(assigned_to="Nobody Real")])
    resp = await client.post(
        "/admin/import/cases",
        headers=_auth(superuser_token),
        files={"file": ("import.xlsx", xlsx, "application/octet-stream")},
    )
    body = resp.json()
    assert body["imported"] == 0
    assert "does not match any existing user" in body["rejected"][0]["reason"]


async def test_import_mixed_valid_and_invalid_rows(
    client: AsyncClient, superuser_token: str, superuser: User, regular_user: User
):
    xlsx = _build_xlsx(
        [
            _valid_row(customer="Valid Row Co"),
            _valid_row(customer=None),
            _valid_row(assigned_to="Ghost User"),
        ]
    )
    resp = await client.post(
        "/admin/import/cases",
        headers=_auth(superuser_token),
        files={"file": ("import.xlsx", xlsx, "application/octet-stream")},
    )
    body = resp.json()
    assert body["totalRows"] == 3
    assert body["imported"] == 1
    assert len(body["rejected"]) == 2
    assert {r["row"] for r in body["rejected"]} == {3, 4}
